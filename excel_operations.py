import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from PyQt5.QtWidgets import QMessageBox
import os

def save_data_to_excel(df, output_file_path):
    try:
        df.to_excel(output_file_path, index=False)
    except Exception as e:
        raise Exception(f"Error saving data to Excel: {e}")

def append_to_final(temp_file_path, final_file_path):
    try:
        temp_df = pd.read_excel(temp_file_path)

        if 'Zone' in temp_df.columns:
            os.remove(temp_file_path)
            raise Exception("Please enter a valid file. The file contains an invalid 'Zone' column.")

        temp_df = temp_df.drop(['Sr.', 'Remarks', 'No. of pieces', 'Weight'], axis=1)

        if not os.path.exists(final_file_path):
            temp_df.to_excel(final_file_path, index=False)
        else:
            final_df = pd.read_excel(final_file_path)
            if temp_df['CN #'].isin(final_df['CN #']).any():
                raise Exception("Tracking ID already exists in the final sheet.")
            final_df = pd.concat([final_df, temp_df], ignore_index=True)
            final_df.to_excel(final_file_path, index=False)

        return True
    except Exception as e:
        QMessageBox.warning(None, "Error", f"Error appending to final sheet: {e}")
        return False

def add_columns(file_path):
    try:
        wb = load_workbook(file_path)
        sheet = wb.active
        headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]

        desired_columns = ["Status", "Recent Location", "COD Amount", "Booking Date", "Payment Received"]

        for column_name in desired_columns[:2]:
            if column_name not in headers:
                sheet.insert_cols(2)
                sheet.cell(row=1, column=2).value = column_name
                headers.insert(1, column_name)

        if "COD Amount" not in headers:
            sheet.insert_cols(len(headers) + 1)
            sheet.cell(row=1, column=len(headers) + 1).value = "COD Amount"
            headers.append("COD Amount")

        for column_name in desired_columns[3:]:
            if column_name not in headers:
                sheet.insert_cols(len(headers) + 1)
                sheet.cell(row=1, column=len(headers) + 1).value = column_name
                headers.append(column_name)

        wb.save(file_path)
    except Exception as e:
        QMessageBox.warning(None, "Error", f"Error adding columns: {e}")

def calculate_payments(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]

    required_columns = ["COD Amount", "Payment Received", "Recent Location"]
    if not all(column in headers for column in required_columns):
        raise ValueError("One or more required columns are missing from the sheet.")

    data = sheet.iter_rows(min_row=2, values_only=True)
    total_cod_amount = 0
    pending_payment = 0
    delivered_pending = 0
    recent_location = None
    returned_statuses = ['Returned to shipper', 'Being Return', 'Pickup Request Sent', 'Ready for Return']

    for row in data:
        cod_amount = row[headers.index("COD Amount")]
        payment_received = row[headers.index("Payment Received")]
        recent_location = row[headers.index("Recent Location")]

        if cod_amount and (payment_received == "" or payment_received == "-") and recent_location not in returned_statuses:
            pending_payment += float(cod_amount)
            if recent_location == "Delivered":
                delivered_pending += float(cod_amount)

        if recent_location not in returned_statuses:
            total_cod_amount += float(cod_amount)

    return total_cod_amount, pending_payment, delivered_pending

def customize_excel(self, directory):
    file_path = os.path.join(directory, 'final.xlsx') if directory else \
        os.path.join(os.environ['USERPROFILE'], 'Desktop', 'final.xlsx')
    
    df = pd.read_excel(file_path)
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    cod_amount_col = df.columns.get_loc('COD Amount') + 1
    status_col = df.columns.get_loc('Status') + 1
    payment_received_col = df.columns.get_loc('Payment Received') + 1
    recent_location_col = df.columns.get_loc('Recent Location') + 1

    for index, row in df.iterrows():
        cod_amount_value = row['COD Amount']
        status_value = str(row['Status']).strip().lower()
        payment_received_value = str(row['Payment Received']).strip().lower()
        recent_location_value = str(row['Recent Location']).strip().lower()

        cod_amount_cell = sheet.cell(row=index + 2, column=cod_amount_col)
        status_cell = sheet.cell(row=index + 2, column=status_col)
        payment_received_cell = sheet.cell(row=index + 2, column=payment_received_col)
        recent_location_cell = sheet.cell(row=index + 2, column=recent_location_col)

        if pd.notna(cod_amount_value) and isinstance(cod_amount_value, (int, float)) and cod_amount_value > 5000:
            cod_amount_cell.fill = yellow_fill

        if 'delivered' in status_value:
            status_cell.fill = green_fill

        if 'ready for return' in recent_location_value or 'return' in recent_location_value:
            recent_location_cell.fill = red_fill

        if 'pending' in recent_location_value:
            recent_location_cell.fill = red_fill

        if 'paid' in payment_received_value:
            payment_received_cell.fill = green_fill
        elif 'pending' in payment_received_value:
            payment_received_cell.fill = yellow_fill
        else:
            payment_received_cell.fill = red_fill

    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[column_letter].width = max_length + 2

    updated_file_path = os.path.join(directory, 'final.xlsx') if directory else \
        os.path.join(os.environ['USERPROFILE'], 'Desktop', 'customized_final.xlsx')
    workbook.save(updated_file_path)
    workbook.close()

def sort_by_booking_date(file_path):
    df = pd.read_excel(file_path)
    df['Booking Date'] = pd.to_datetime(df['Booking Date'], format='%d/%m/%Y', errors='coerce')
    sorted_df = df.sort_values(by='Booking Date', ascending=True)
    sorted_df['Booking Date'] = sorted_df['Booking Date'].dt.strftime('%d/%m/%Y')
    sorted_df.to_excel(file_path, index=False)

def calculate_pending_count(file_path):
    df = pd.read_excel(file_path)
    if 'Payment Received' not in df.columns:
        raise Exception("The 'Payment Received' column is not found in the sheet.")
    pending_count = (df['Payment Received'] == '-').sum()
    return pending_count
