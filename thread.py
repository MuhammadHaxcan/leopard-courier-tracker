from PyQt5.QtCore import QThread, pyqtSignal
from openpyxl import load_workbook

class WorkerThread(QThread):
    progress = pyqtSignal(int)
    result = pyqtSignal(str)
    error = pyqtSignal(str)

    def __init__(self, api, final_file_path, mode):
        super().__init__()
        self.api = api
        self.final_file_path = final_file_path
        self.mode = mode

    def run_tracking(self):
        try:
            wb = load_workbook(self.final_file_path)
            sheet = wb.active

            if sheet.max_column < 3:
                sheet.insert_cols(2, 2)
                sheet.cell(row=1, column=2).value = "Status"
                sheet.cell(row=1, column=3).value = "Recent Location"

            total_rows = sheet.max_row - 1
            if total_rows <= 0:
                self.error.emit("No data to process.")
                return

            progress_interval = max(1, total_rows // 100)

            for row in range(2, sheet.max_row + 1):
                track_number = sheet.cell(row=row, column=1).value
                current_status = sheet.cell(row=row, column=2).value

                if current_status and current_status.lower() == "delivered":
                    continue

                if track_number:
                    try:
                        booked_packet_status, recent_status, booking_date = self.api.track_booked_packet(track_number)

                        if booked_packet_status:
                            sheet.cell(row=row, column=2).value = booked_packet_status
                        if recent_status:
                            sheet.cell(row=row, column=3).value = recent_status
                        if booking_date:
                            sheet.cell(row=row, column=sheet.max_column - 1).value = booking_date

                    except Exception as api_error:
                        self.error.emit(f"Failed to track packet {track_number}: {api_error}")
                        continue
                else:
                    self.error.emit(f"Track number missing in row {row}. Skipping...")

                if (row - 1) % progress_interval == 0 or row == sheet.max_row:
                    progress = int((row - 1) / total_rows * 100)
                    self.progress.emit(progress)

            try:
                wb.save(self.final_file_path)
            except Exception as save_error:
                self.error.emit(f"Failed to save the file: {save_error}")
                return

            self.result.emit(f"Status update completed. Data has been updated.")
            self.progress.emit(100)

        except Exception as e:
            self.error.emit(f"An unexpected error occurred: {e}")

    def run_payment(self):
        try:
            wb = load_workbook(self.final_file_path)
            sheet = wb.active

            if sheet.max_column < 10:
                sheet.insert_cols(10)
                sheet.cell(row=1, column=10).value = "Payment Status"

            total_rows = sheet.max_row - 1
            if total_rows <= 0:
                self.error.emit("No data to process.")
                return

            progress_interval = max(1, total_rows // 100)

            tracking_ids = []
            row_mapping = {}

            for row in range(2, sheet.max_row + 1):
                track_number = sheet.cell(row=row, column=1).value
                current_payment_status = sheet.cell(row=row, column=10).value

                if current_payment_status and current_payment_status.lower() == 'paid':
                    continue

                if track_number:
                    tracking_ids.append(track_number)
                    row_mapping[track_number] = row

                if len(tracking_ids) == 50 or (row == sheet.max_row and tracking_ids):
                    tracking_ids_str = ','.join(tracking_ids)
                    try:
                        parsed_results = self.api.track_payment_status(tracking_ids_str)

                        for track_id, (payment_status, payment_date) in parsed_results.items():
                            if track_id in row_mapping:
                                if (payment_status == None):
                                    recent_status = '-'
                                else:
                                    recent_status = f'{payment_status} {payment_date}'.strip()
                                sheet.cell(row=row_mapping[track_id], column=10).value = recent_status

                        tracking_ids.clear()
                        row_mapping.clear()

                    except Exception as api_error:
                        self.error.emit(f"Failed to track payment for batch starting at row {row - len(tracking_ids) + 1}: {api_error}")
                        tracking_ids.clear()
                        row_mapping.clear()
                        continue

                if (row - 1) % progress_interval == 0 or row == sheet.max_row:
                    progress = int((row - 1) / total_rows * 100)
                    self.progress.emit(progress)

            try:
                wb.save(self.final_file_path)
            except Exception as save_error:
                self.error.emit(f"Failed to save the file: {save_error}")
                return

            self.result.emit("Payment tracking completed. Data has been updated.")
            self.progress.emit(100)

        except Exception as e:
            self.error.emit(f"An unexpected error occurred: {e}")

    def run(self):
        if self.mode == "tracking":
            self.run_tracking()
        elif self.mode == "payment":
            self.run_payment()
        else:
            self.error.emit("Invalid operation mode selected.")
